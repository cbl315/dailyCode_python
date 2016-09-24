#!/usr/bin/env python
# 第 0014 题： 纯文本文件 student.txt为学生信息, 里面的内容（包括花括号）如下所示：
'''
{
    "1":["张三",150,120,100],
    "2":["李四",90,99,95],
    "3":["王五",60,66,68]
}
'''
import json
import openpyxl
import collections
import os


def loadTxtInfoOfJson(txt_path):
    f = open(txt_path, 'r')
    # print(type(f))--> <class '_io.TextIOWrapper'>
    content = json.load(f)
    # print(content, type(content))--> type=<class 'dict'>
    return content


def writeDictToExcel(d, sheet_name):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    l = sorted(d)
    for index in range(len(l)):
        # 字典的key值存到excel的第一列（openxypl的row和column从1开始计数）
        ws.cell(row=index+1, column=1, value=l[index])
        # 默认value值对应的是list、tuple等iterable的数据结构
        if isinstance(d[l[index]], collections.Iterable):
            for i in range(len(d[l[index]])):
                ws.cell(row=index+1, column=2+i, value=d[l[index]][i])
        else:
            ws.cell(row=index+1, column=2, value=d[l[index]])
    wb.save('student.xlxs')

if __name__ == '__main__':
    txt_path = r'student.txt'
    sheet_name = os.path.basename(txt_path)
    sheet_name = sheet_name[0:len(sheet_name) - 4]
    json_dict = loadTxtInfoOfJson(txt_path)
    writeDictToExcel(json_dict, sheet_name)
    print('success!')

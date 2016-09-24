#!/usr/bin/env python
import openpyxl
import os
from lxml import etree


def transform(xml_path):
    f = open(xml_path, 'r', encoding='utf-8')
    text = f.read()
    f.close()
    text = text.replace(r'&gt;', '>')
    text = text.replace(r'&lt;', '<')
    f = open(xml_path, 'w', encoding='utf-8')
    f.write(text)
    f.close()


def getGeneratorLen(g):
    # 超过 100,000则返回 -1
    count = 0
    for i in g:
        if count < 100000:
            count += 1
        else:
            return -1
    return count


def readXls(xls_path):
    wb = openpyxl.load_workbook(xls_path)
    file_name = os.path.basename(xls_path).split('.')
    file_name = '.'.join(file_name[0:len(file_name)-1])
    sheet = wb.get_sheet_by_name(file_name)
    return sheet


def sheetToStr(sheet):
    result = '''{'''
    for r in range(1, getGeneratorLen(sheet.rows) + 1):
        tmp = ''
        for col in range(1, getGeneratorLen(sheet.columns) + 1):
            if sheet.cell(row=r, column=col).value is not None:
                if col == 1:
                    tmp += ('[%s,' %
                            sheet.cell(row=r, column=col).value)
                else:
                    if type(sheet.cell(row=r, column=col).value) == str:
                        tmp += '"%s",' % (
                            sheet.cell(row=r, column=col).value)
                    else:
                        tmp += '%s,' % (
                            sheet.cell(row=r, column=col).value)
        tmp = tmp[0:len(tmp)-1] + '],'
        # print(tmp)
        result += '\n\t%s' % tmp
    result = result[0:len(result)-1] + '\n}'
    return result


def buildXml(sheet, file_name):
    root = etree.Element('root')
    result_str = sheetToStr(sheet)
    root.text = '\n'
    result_xml = etree.ElementTree(root)
    comment = """
<!--
    数字信息
-->
"""
    students = etree.SubElement(root, file_name)
    students.text = comment + '%s\n' % result_str
    students.tail = '\n'
    f = open('0019.xml', 'w', encoding='utf-8')
    f.write('<?xml version="1.0" encoding="UTF-8"?>\n')
    f.write(etree.tounicode(result_xml.getroot()))
    f.close()
    transform('0019.xml')


if __name__ == '__main__':
    xls_path = r'numbers.xlsx'
    student = readXls(xls_path)
    file_name = os.path.basename(xls_path).split('.')
    file_name = '.'.join(file_name[0:len(file_name)-1])
    buildXml(student, file_name)
    print('success!')

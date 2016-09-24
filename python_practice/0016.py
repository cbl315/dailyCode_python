#!/usr/bin/env python
# 第 0016 题： 纯文本文件 numbers.txt, 里面的内容（包括方括号）如下所示：
'''
[
    [1, 82, 65535],
    [20, 90, 13],
    [26, 809, 1024]
]
请将上述内容写到 numbers.xls 文件中
'''
import json
import openpyxl
import collections
import os


def loadTxtInfoOfJson(txt_path):
    f = open(txt_path, 'r')
    # print(type(f))--> <class '_io.TextIOWrapper'>
    content = json.load(f)
    # print(content, type(content))--> type=<class 'dict'>
    return content


def writeDictToExcel(d, sheet_name):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    l = sorted(d)
    for index in range(len(l)):
        # 字典的key值存到excel的第一列（openxypl的row和column从1开始计数）
        ws.cell(row=index+1, column=1, value=l[index])
        # 默认value值对应的是list、tuple等iterable的数据结构
        if (isinstance(d[l[index]], collections.Iterable) and
                type(d[l[index]]) != str):
            for i in range(len(d[l[index]])):
                ws.cell(row=index+1, column=2+i, value=d[l[index]][i])
        else:
            ws.cell(row=index+1, column=2, value=d[l[index]])
    wb.save('%s.xlxs' % sheet_name)


def writeListToExcel(l, sheet_name):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    for index in range(len(l)):
        if (isinstance([l[index]], collections.Iterable) and
                type([l[index]]) != str):
            for i in range(len(l[index])):
                ws.cell(row=index+1, column=1+i, value=l[index][i])
        else:
            ws.cell(row=index+1, column=1, value=[l[index]])
    wb.save('%s.xlxs' % sheet_name)
if __name__ == '__main__':
    txt_path = r'numbers.txt'
    sheet_name = os.path.basename(txt_path)
    sheet_name = sheet_name[0:len(sheet_name) - 4]
    json_dict = loadTxtInfoOfJson(txt_path)
    if type(json_dict) == dict:
        writeDictToExcel(json_dict, sheet_name)
    elif(type(json_dict)) == list:
        writeListToExcel(json_dict, sheet_name)
    print('success!')
